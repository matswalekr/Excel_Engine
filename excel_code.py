import os
import pandas as pd
from openpyxl import load_workbook, Workbook
import openpyxl as openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from pycel import ExcelCompiler
from openpyxl.utils.cell import coordinate_from_string, coordinate_to_tuple

from typing import Any, Generator, Union, Tuple, Literal, Dict, Self, List
from dataclasses import dataclass, field
import warnings
import operator

# Suppress DeprecationWarnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

class DocstringInheritor(type):
    '''
    Metaclass for docstring inheritance for both the main class and its methods.
    '''
    def __new__(cls, name, bases, dct):
        new_cls = super().__new__(cls, name, bases, dct)

        # Inherit class docstring
        if not new_cls.__doc__:
            for base in bases:
                if base.__doc__:
                    new_cls.__doc__ = base.__doc__
                    break

        # Inherit method docstrings
        for attr_name, attr in dct.items():
            if not attr.__doc__:
                for base in bases:
                    base_attr = getattr(base, attr_name, None)
                    if base_attr and base_attr.__doc__:
                        attr.__doc__ = base_attr.__doc__
                        break

        return new_cls


class AutoNumericOperations(type):

    """
    Metaclass to automate mathematical operations (add, iadd, etc)
    Written by Chat-GPT, dont blame me"""

    def __new__(cls, name, bases, dct):
        # Define operations dynamically, adding custom error handling as in your methods
        for op in ['add', 'sub', 'mul', 'truediv', 'floordiv', 'mod', 'pow']:
            dct[f'__{op}__'] = cls._create_bin_op(op)

        for op in ['iadd', 'isub', 'imul', 'itruediv', 'ifloordiv', 'imod', 'ipow']:
            dct[f'__{op}__'] = cls._create_inplace_op(op)
        
        return super().__new__(cls, name, bases, dct)

    @staticmethod
    def _create_bin_op(op):
        """Creates a binary operation method like __add__, __sub__ etc."""
        def bin_op(self, other: Any):
            # Default behavior for when other is a Cell or another type
            if isinstance(other, self.__class__):
                other = other.value
            try:
                return getattr(operator, op)(self.value, other)
            except ValueError:
                raise ValueError(f"Trying to {op} type {type(other)} to {self.__class__.__name__} with data of type {type(self.value)}")
        return bin_op

    @staticmethod
    def _create_inplace_op(op):
        """Creates an in-place operation method like __iadd__, __isub__ etc."""
        def inplace_op(self, other: Any):
            # Default behavior for when other is a Cell or another type
            if isinstance(other, self.__class__):
                other = other.value
            try:
                setattr(self, 'value', getattr(operator, op)(self.value, other))
                return self
            except ValueError:
                raise ValueError(f"Trying to {op} type {type(other)} to {self.__class__.__name__} with data of type {type(self.value)}")
        return inplace_op


class Excel(metaclass = DocstringInheritor):
    """"
    Class representing an Excel file.

    Allows you to manipulate cells and functions easily.


    path: str
    __________________________________

    References the path of the Excel file.

    If no path is given, an empty Excel file is created in 'w' or 'm' mode.

    If the Excel class is opened in 'w' mode, a non-existing path will create an empty Excel workbook under that path.


    Optional: open_path: bool = False
    __________________________________

    Optional args to immediately open the path rather than with the __enter__ method.

    For best pratices use: with Excel(path) as doc:
        ...


    Some functionality might not be accessible using the class itself. Its attribute worksheet represents an openpyxl/pycel Workbook object, which can be manipulated.

    In addition, the sheet can be accessed using self.pd as a pandas.DataFrame. Note that it is not automatically saved, but needs to be saved using self.save(df = 'Dataframe')

    Has a Cell subclass representing a cell.
    """

    @dataclass(slots=True, unsafe_hash=True, eq = True, repr=True)
    class Cell(metaclass = AutoNumericOperations):
        '''
        Dataclass for Cells in Excel files. Usage of __slots__ (in decorator) optimises memory usage by only allowing the pre-defined attributes for this class.

        All comparison operators are implemented, but only compare position in a sheet, not the cell's value.

        Rows and columns can be incremented using shift_columns/shift_cells (Note that this does not yet update the cells in the workbook)
        <, <=, >, >= compare by position in a sheet. The further left, then up a cell exists, the lower it is. If sheets are different, an error is raised.

        If the value of the cell needs to be compared, use Cell.value

        == also compares the sheet in addition to row and column. It does not compare the value of the cells.

        In-place operations are made using the values. The Cell address stays the same
        '''
        value_: Any = field(default=None, hash=False, repr=True, compare=False)
        col: str = field(default=None, hash=True, repr=True, compare=True)
        row: int = field(default=None, hash=True, repr=True, compare=True)
        sheet_name: str = field(default=None, hash=True, repr=True, compare=True)
        workbook: Workbook = field(default=None, hash=True, repr=True, compare=True)
        name: str = field(default= None, hash=False, repr=False, compare=False)
        filename: str = field(default= None, hash=False, repr=False, compare=True)
        excel_compiler:ExcelCompiler = field(default= None, hash=False, repr=False, compare=True)
        sheet: openpyxl.worksheet = field(default= None, hash=False, repr=False, compare=True)


        def __post_init__(self)-> None:
        # Make sure we have initialised all variables necessary
            
            if not isinstance(self.row, int):
                self.row = int(self.row)

            if self.filename:
                if not self.excel_compiler:
                    self.excel_compiler = ExcelCompiler(filename=self.filename)
                if not self.workbook:
                    self.workbook = load_workbook(filename=self.filename)

            if not self.sheet:
                self.sheet = self.workbook[self.sheet_name]
            if self.value_:
                self.value = self.value_

        @property
        def pos(self)->str:
            """
            Returns a string representing the excel position of a Cell
            """
            return f"{self.col}{self.row}"
            
        @property
        def value(self)-> Union[str,int,float]:
            """
            Automatically evaluates the cell's value if the Cell has a formula. 
            Else, it returns the value with which the cell was initialised"""
            if isinstance(self.value_, str) and "=" in self.value_:
                print("= found")
                compiled_value:Any = self.excel_compiler.evaluate(f"{self.sheet_name}!{self.pos}")
                try:
                    return float(compiled_value)
                except TypeError:
                    return compiled_value
            else:
                return self.sheet.cell(row = self.row, column = column_index_from_string(self.col)).value
            
        @value.setter
        def value(self, new_value: Union[str, int, float]):
            self.sheet.cell(row = self.row, column = column_index_from_string(self.col)).value = new_value
            self.value_ = new_value
    
        def __str__(self)->str:
            return str(self.value)
        
        def __le__(self, other:Self)->bool:
            return any([self == other, self < other])
        
        def __ge__(self, other:Self)->bool:
            return any([self==other, self>other])
        
        def __gt__(self, other:Self)->bool:
            if not isinstance(other, Excel.Cell):
                raise ValueError(f"The compared object is not of type Excel.Cell, but of type {type(other).__name__}")
            else:
                if not (self.sheet == other.sheet):
                    raise ValueError(f"The compared cell is not in the same sheet: Sheet = {self.sheet} vs Sheet{other.sheet}")
                else:
                    if self.col == other.col:
                        return int(self.row) > int(other.row)
                    elif len(self.col) != len(other.col):
                        return len(self.col) > len(other.col)
                    else:
                        return self.col > other.col
        
        def shift_column(self, shift: int) -> None:
            if not self.col:
                raise ValueError("Column name cannot be empty.")

            try:
                col_num = column_index_from_string(self.col)  # Convert column name to numeric index
            except ValueError:
                raise ValueError(f"Invalid column name: {self.col}")

            new_col_num = col_num + shift
            if new_col_num < 1:
                raise ValueError("Column number out of range.")

            self.col = get_column_letter(new_col_num)  # Convert numeric index back to column name

        def shift_row(self, shift: int)-> None:
            if not self.row:
                raise ValueError("No row initialised")
            
            new_row = self.row + shift

            if new_row < 1:
                raise ValueError("Row number out of bounds")
            self.row = new_row
            
    def __init__(self, path:str, open_path:bool=False)->None:
        if path:
            self.path = os.path.abspath(path)
        else:
            self.path = None

        if open_path:
            warnings.warn("\nIt is prefered to open Excel files using: \nwith Excel(path) as doc:\n   ....\nor the open_excel(path:str, mode: ['r', 'm', 'w']) function", UserWarning)
            self.open_doc()

        self.df_: Dict[str, pd.DataFrame] = None
        self.sheets_:list = None
        self.active_ = None #Active sheet
        self.defined_names_ = None #Defined names of the Cells
        self.path_intermediate_save= r"intermediate_save.xlsx"

    @property
    def sheets(self)->List[str]:
        """
        Returns a list of all the names of the sheets of an excel file"""

        return self.workbook.sheetnames
    
    @property
    def active(self)-> openpyxl.worksheet:
        '''
        Returns the active worksheet of the excel file
        '''
        return self.workbook.active

    @property
    def df(self)->Dict[str,pd.DataFrame]:

        '''
        Returns a dictionary of str: pd.DataFrame of the excel document.

        The string key is the name of the worksheet
        '''
        pd.read_excel(path = self.path_intermediate_save)
        dfs = {}

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            data = sheet.values

            try:                                                            # Check if the sheet is empty
                columns = next(data)
            except StopIteration:
                dfs[sheet_name] = pd.DataFrame()                            # If the sheet is empty, create an empty DataFrame
                continue

            df = pd.DataFrame(data, columns = columns)
            dfs[sheet_name] = df
        return dfs
    
    @property
    def defined_names(self)->Dict[str, str]:
        """
        Returns all defined names of the spreadsheet"""
        return self.workbook.defined_names
    
    @staticmethod
    def split_cell(cell_name:str)->Tuple[str, int]:
        '''
        Splits a string referencing a cell into its column and row and returns them as str and int respectively
        Returns column, row
        '''
        return coordinate_from_string(cell_name)

    @staticmethod
    def column_str_to_int(column:str)->int:
        """
        Transforms the string representation of a column in excel into its int representation
        Used to interface with openpyxd
        """
        return int(column_index_from_string(column))
    
    @staticmethod
    def column_int_to_str(column: int) -> str:
        """
        Transforms the string representation of a column in excel into its int representation
        Used to interface with openpyxd
        """
        return str(get_column_letter(column))

    @staticmethod
    def split_cell_int(cell:str) -> Tuple[int, int]:
        """
        Splits a string referencing a cell into ist column and row and returns them as ints
        Used to interface with openpyxl
        Returns column, row as integers
        """
        column, row = Excel.split_cell(cell)

        return int(Excel.column_str_to_int(column)), row

    @staticmethod
    def char_range(start,end)->Generator[str, None, None]:
        """
        Generate a range of characters from start to end, inclusive.
        """
        for char_code in range(ord(start), ord(end) + 1):
            yield chr(char_code)

    @staticmethod
    def _copy_sheet_contents(target_sheet, source_sheet)-> None:
        '''
        Copys the contents of one excel sheet into a different one
        '''
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    # Copy each style attribute individually
                    new_cell.font = cell.font
                    new_cell.fill = cell.fill
                    new_cell.border = cell.border
                    new_cell.alignment = cell.alignment
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection
                if cell.hyperlink:
                    new_cell.hyperlink = cell.hyperlink
                if cell.comment:
                    new_cell.comment = cell.comment

    def get_address_named_cell(self, name:str)->Tuple[str, str]:
        """
        Returns sheet_name, cell_address"""
        if name not in self.defined_names:
            raise ValueError(f"Named cell {name} not found")
        
        defined_name = self.workbook.defined_names[name]

        destinations = list(defined_name.destinations)
        if not destinations:
            raise ValueError(f"Named cell {name} does not have any associated destinations")

        if len(destinations) > 1:
            # Either raise an error or decide how to handle multiple cells
            raise ValueError(f"Named cell {name} refers to multiple destinations")

        sheet_name, cell = destinations[0]
        return (sheet_name, cell)

    def __enter__(self)->Self:
        self.open_doc()
        return self
    
    def __exit__(self, exc_type, exc_value, traceback)->None:
        del self.workbook
        del self

    def __len__(self)->int:
        """
        Returns the number of sheets in the doc"""
        return len(self.sheets)
    
    def __repr__(self)->str:
        return f"Excel(path={self.path})"
    
    def __str__(self)-> str:

        str_return = '\n'

        width = 24

        for key, value in self.df.items():

            str_return += f"{key:^{width}}"
            str_return += f"\n{'_' * width}\n\n"
            if not value.empty:
                str_return += value.to_string() + '\n\n'
            else: str_return += '\n\n'

        return str_return

    def open_doc(self, intermediate = False, read_only:bool = False)->None:
        '''
        Function used to open an Excel document.

        Called by the __enter__ method and, if specified, the __init__ method.
        '''
        if not intermediate:
            path_open = self.path

        else:
            path_open = self.path_intermediate_save

        if read_only:
            self.workbook: Workbook = load_workbook(path_open, 
                                                data_only = False, 
                                                keep_vba=True, 
                                                keep_links=True,
                                                read_only=True)
        
        else:
            self.workbook: Workbook = load_workbook(path_open, 
                                                data_only = False, 
                                                keep_vba=True, 
                                                keep_links=True)
          
    def copy(self, from_intermediate = True)->Self:
        if from_intermediate:
            path = self.path_intermediate_save
        else:
            path = self.path
        
        return type(self)(path = path, open_path = True)

    def open_excel(self)->None:
        """
        Opens the file in Excel itself"""
        self.save(intermediate = False)
        os.startfile(self.path)

    def range(self, range_:str = False, sheet:str = None, raw:bool = False, value_only:bool = False, pos_only:bool=True, cell:bool= False)-> Generator[Union[Cell, Any, str], str, None]:
        '''
        Returns an iterator of the individual cells in a range (ex. A1:A3 -> returns cells A1, A2, A3)

        Return type is a string of position.

        if raw = True, a Cell object is returned where the value of the Cell is a cell object of aps.

        if value_only = True, only the value of the cells is returned

        if cell = True, a Cell object is returned where value is the value of the cell
        '''
        if not sheet: sheet = self.active
        else: sheet = self.workbook[sheet]

        if range_:
            start, end = range_.split(':')
            column_start, row_start = Excel.split_cell(start)
            column_end, row_end = Excel.split_cell(end)
        else:
            column_start, row_start = 'A', 1
            column_end, row_end = [chr(i) for i in range(ord('A'), ord('Z') + 1)][sheet.max_column - 1], sheet.max_row

        for col in Excel.char_range(column_start, column_end):
            for row in range(row_start, row_end + 1):
                pos: str = col + str(row)
                if raw:
                    yield self.Cell(sheet[pos], col, row, sheet)
                elif value_only:
                    yield sheet[pos].value
                elif cell:
                    yield self.Cell(value_ = sheet[pos].value, col = col, row = row, sheet = sheet)
                elif pos_only:
                    yield pos
                else:
                    raise ValueError('Nothing is specified to be returned')

    def iter_rows(self, min_row:int = 1, max_row:int = None, min_col:int = 1, max_col:int = None, sheet:str = None)-> Generator[Any, None, None]:
        '''
        returns an iterator going over all cells in a certain range in a sheet
        '''
        if not sheet: sheet = self.active
        if not max_row: max_row = sheet.max_row
        if not max_col: max_col = sheet.max_column

        return sheet.iter_rows(min_row = min_row, max_row = max_row, min_col = min_col, max_col = max_col)

    def save(self, intermediate = False, path:str = None, df:pd.DataFrame = None, header:bool = False)-> None:
        '''
        By default, this saves the file in its original path.

        If necessary, a new path can be specified.

        It is also possible to save a pd.DataFrame using this method,
        in case the excel file has been exported as a df and then be manipulated in any shape or form. Note that it will open the new Excel document as the self.workbook.
        '''
        if path:
            pass
        elif intermediate:
            path = self.path_intermediate_save
        elif self.path:
            path = self.path
        else:
            raise ValueError('No path was found to save the excel sheet')

        if df is not None:                                                                      # Saves a df object in the specified path
            df.to_excel(path, index = False, header = header,  engine = 'openpyxl')             #, sheet_name = sheet_name)
            self.workbook = load_workbook(path)

        else: 
            self.workbook.save(path)                                                          # Saves the workbook (default)


class Excel_read(Excel):
    """
    Excel object that allows for reading the contents of the file, but not any operations updating its values"""

    def __getitem__(self,index:str)->Union[Excel.Cell, openpyxl.worksheet]:
        '''
        Enables Excel[XXX] behaviour to access and return values of the file.

        Written to support [spreadsheet name], [name cell] and [cell adress]

        Does not support range of cells
        '''

        # Case where workbook
        if index in self.sheets:
            return self.workbook[index]
        
        # Case where Cell
        else:
            if index in self.defined_names:
                for sheet_name, cell in self.defined_names[index].destinations:
                    sheet = self.workbook[sheet_name]
                    column, row = coordinate_from_string(cell)

            else:
                column, row = self.split_cell(index)
                sheet = self.active

            return_cell = self.Cell(workbook=self.workbook, sheet=sheet, col = column, row = row)
            return return_cell

    def __contains__(self, item: Union[str, Excel.Cell])->bool:
        '''
        Implemented for a str representing a sheet and self.Cell representing a cell object
        ''' 
        if isinstance(item, str):
            return item in self.sheets

        if isinstance(item, Excel.Cell):
            return self[item.sheet_name][item.pos].value == item.value

        return False
    
    def name(self, range_:str, name:str, sheet:str = None)-> None:
        '''
        Allows ot give a cell or a range of cells a name to find them easier in future
        '''
        if sheet:
            sheet = self.workbook[sheet]
        else:
            sheet = self.active

        self.workbook.create_named_range(name, sheet, range_)

    def open_doc(self, intermediate=False, read_only:bool = True) -> None:
        return super().open_doc(intermediate, read_only = read_only)
    

class Excel_modify(Excel_read):
    """
    Class of the excel sheet that allows modification of the data while it is open.
    It cant be saved however"""
    
    def __init__(self, path:str)->Self:
        super().__init__(path, open_path=False)

    @property
    def pycel(self)->None:
        return ExcelCompiler(filename=self.path)
    
    def evaluate_formula(self, cell: Union[Excel.Cell,str]) -> Any:
        """
        Evaluates a formula using the excel sheet.
        Input can be the Cell itself or a string pointing to the cell such as "D12"
        """
        if isinstance (cell, self.Cell):
            cell = cell.pos
        
        return self.pycel.evaluate(f"{self.active.title}!{cell}")
    
    def __setitem__(self, index, value:Any)->None:
        """
        Allows setting the value of cells or sheets using [] notation.

        Supports [sheet name], [named range], [cell range], [named cell], and [cell].

        Supports cell ranges using examples like 'A2:A23'.
        """
        # Case when the index is a sheet
        if index in self.sheets:
            self.workbook[index] = value

        # Else we are assigning a value to a cell or range of cells
        else:
            if isinstance(value, Excel.Cell):
                value = value.value

            # Case where we have a named range
            if ":" in index:
                start, end = index.split(':')
                column_start, row_start = Excel.split_cell(start)
                column_end, row_end = Excel.split_cell(end)

                for col in Excel.char_range(column_start, column_end):
                    for row in range(row_start, row_end+1):
                        pos:str = f"{col}{row}"
                        print(pos)
                        self[pos] = value

            # Else we must be assigning a value to a cell
            else:
                if isinstance(value, Excel.Cell):
                    value = value.value
                
                # Case when we used a defined name
                if index in self.defined_names:
                    sheet_name, cell = self.get_address_named_cell(index)
                    col, row = Excel.split_cell_int(cell)
                    self.workbook[sheet_name].cell(row = row, column = col).value = value
                else:
                    cell = index
                    col, row = Excel.split_cell_int(index)
                    self.active.cell(row= row, column = col).value = value

    def __delitem__(self, index)-> None:
        '''
        Allows to delete a sheet, named cells or a range of cells.

        Note that for ease, deleting means setting the value of a a cell to 0.
        '''
        if index in self.sheets:
            self.workbook.remove(index)

        elif index in self.defined_names:
            for sheet_name, cell in self.defined_names[index].destinations:
                self[sheet_name][cell].value = 0

        elif ':' in index:
            start, end = index.split(':')
            column_start, row_start = Excel.split_cell(start)
            column_end, row_end = Excel.split_cell(end)

            for col in Excel.char_range(column_start, column_end):
                for row in range(row_start, row_end + 1):
                    pos: str = col + str(row)
                    self.active[pos] = 0

        else:
            self.workbook.active[index] = 0        
    
    def new_sheet(self, name:str)-> None:
        '''
        Creates a new sheet in an excel file
        '''
        if name not in self.sheets:
            self.workbook.create_sheet(title = name)

    def rename_sheet(self, name_old:str, name_new:str)-> None:
        '''
        Renames a sheet of the excel file
        '''
        self.workbook[name_old].title = name_new

    def vba(self, name:str = None, vba_code:str = None, vba_name:bool = False, file_name:str = None)-> None:
        '''
        Allows to run saved VBAs or to insert a VBA code to be run on the file
        '''
        assert any([name, vba_code]), 'No VBA given'

        workbook = self.excel.Workbooks.Open(self.path)

        if name: self.excel.Application.Run(name)
        if vba_code:
            vb_project = self.workbook.VBProject
            vb_module = vb_project.VBComponents.Add(1)
            vb_module.CodeModule.AddFromString(vba_code)

            def get_subroutine_names(vb_module):
                code_lines = vb_module.CodeModule.Lines(1, vb_module.CodeModule.CountOfLines)
                lines = code_lines.split('\n')
                subroutine_names = []
                for line in lines:
                    line = line.strip()
                    if line.lower().startswith('sub '):
                        sub_name = line.split(' ')[1].split('(')[0]
                        subroutine_names.append(sub_name)
                return subroutine_names

        # Get the subroutine names
        subroutine_names = get_subroutine_names(vb_module)

        # Run each subroutine found
        for subroutine in subroutine_names:
            try:
                self.excel.Application.Run(subroutine)
            except Exception as e:
                print(f"Error running subroutine {subroutine}: {e}")

    def clear_sheet(self, sheet, range_: str = None)-> None:
        '''
        Clears an entire sheet by default.

        If needed, a range to clear can be given. In this case only that range will be cleared.
        '''
        for pos in self.range(range_ = range_, sheet = sheet):
            self[sheet][pos] = None

    def clear_workbook(self, sheets:list[str] = None)-> None:
        '''
        Clears the entire workbook by default.

        If requried, only a sheet/range to clear can be given
        '''
        if not sheets:
            all_sheets = True
            sheets = self.sheets

        for sheet in sheets: self.workbook.remove(self.workbook[sheet])

        if all_sheets:
            self.new_sheet('Sheet1')                                                             # Creates on starting sheet in case all other sheets are deleted. One sheet is always needed in a workbook for openpyxl funcionality

    def remove_sheet(self, sheet_name: str)-> None:
        self.workbook.remove(self.workbook[sheet_name])

    def set_cells_pandas(self, start_cell: str, df: pd.DataFrame, sheet_name:str = None, index:bool = False)->None:
        """
        Allows you to set a range of cells to the contents of a pandas data frame"""

        if start_cell in self.defined_names:
            sheet_name, start_cell = self.get_address_named_cell(name = start_cell)

        if not sheet_name:
            sheet_name = self.active


        start_row, start_col = coordinate_to_tuple(start_cell)
        number_rows: int = df.shape[0]
        number_columns: int = df.shape[1]

        if index:
            df.reset_index(inplace = True)
            number_columns += 1

        for i in range(0, number_rows):
            for j in range(0, number_columns):
                self.workbook[sheet_name].cell(row=start_row + i, column=start_col + j, value=df.iloc[i,j])
        #for i, row in df.iterrows():
            #for j, value in enumerate(row):
                # Write the value to the corresponding cell
                #self.workbook[sheet_name].cell(row=start_row + i, column=start_col + j, value=value)

    def open_doc(self, intermediate=False) -> None:
        return super().open_doc(intermediate = intermediate, read_only = False)
            

class Excel_write(Excel_modify):
    
    def safety_save(self)-> None:
        '''
        Saves a deepcopy of the current workbook in self.safety_copy.
        '''
        self.safety_copy = self.save(path = self.path_intermediate_save)

    def save(self, path:str = None, df:pd.DataFrame = None, header:bool = False)-> None:
        '''
        By default, this saves the file in its original path.

        If necessary, a new path can be specified.

        It is also possible to save a pd.DataFrame using this method,
        in case the excel file has been exported as a df and then be manipulated in any shape or form. Note that it will open the new Excel document as the self.workbook.
        '''
        if not path:
            if self.path:
                path = self.path
            else:
                raise ValueError('No path was found to save the excel sheet')

        if df is not None:                                                                      # Saves a df object in the specified path
            df.to_excel(path, index = False, header = header,  engine = 'openpyxl')             #, sheet_name = sheet_name)
            self.workbook = load_workbook(path)

        else: self.workbook.save(path)                                                          

    def vba(self, name:str = None, vba_code:str = None, vba_name:bool = False, file_name:str = None)-> None:

        super().vba(name, vba_code, vba_name, file_name)
        self.workbook.Save()
        self.workbook.Close()

    def open_doc(self)-> None:
        '''
        Function used to open an Excel document.

        Called by the __enter__ method and, if specified, the __init__ method.

        If no path is given, an empty Excel workbook is created.

        If the path does not exist yet, it is populated with an empty Excel workbook.
        '''
        try:
           super().open_doc()

        except FileNotFoundError:
            if 'Y' == input(f'Do you want to create a new Excel workbook in the path {self.path} (Y/n)? '):
                self.workbook: Workbook = Workbook()
                self.save()
            else: self = None
    
    def __exit__(self, exc_type, exc_value, traceback)-> None:

        if exc_type is None:
            self.save()

        super().__exit__(exc_type, exc_value, traceback)


def open_excel(path:str, mode: Literal['r', 'm', 'w'])-> Union[Excel_read, Excel_modify, Excel_write]:
    '''
    Returns an instance of an Excel class with the appropriate methods.

    path:str
    __________________________________

    Specifies the path under which the document is saved on the computer.
    Note that it needs to be in the rawstring format (r'').

    If there is no Excel document yet in this path, the path is populated with an empty Excel document.



    mode:str
    __________________________________

    Case 'r': The Excel document is in reading mode. No modifications to the Excel object can be made.

    Case 'm': The Excel document is in modifying mode. Modifications can be made to the Excel document, but it can not be saved.

    Case 'w': The Excel document is in writting mode. Modifications can be made to the Excel document and it can be saved in the same or a different path.
    '''
    match mode:

        case 'r': return Excel_read(path)

        case 'm': return Excel_modify(path)

        case 'w': return Excel_write(path)

        case _:   raise ValueError(f'Mode {mode} is not definde')

